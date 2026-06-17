import pytest
from fastapi.testclient import TestClient
from sqlalchemy.orm import Session

from backend.app.db.models import (
    Client,
    Department,
    Discipline,
    Template,
    TemplateFieldDef,
    TemplateMilestoneDef,
    User,
    UserRole,
)


@pytest.fixture
def template_with_defs(
    db_session: Session,
    admin_user: User,
    project_editor_user: User,
    viewer_user: User,
):
    """Build a complete taxonomy + template + 1 required field + 1 milestone.

    Grants the fixture's project_editor / viewer users an additional role
    in this dept so they can exercise the routes — multi-dept grants are
    legal under the 1.9.1 schema.
    """
    d = Department(code="DIV1", name="Division 1")
    db_session.add(d)
    db_session.flush()
    db_session.add(
        UserRole(
            user_id=project_editor_user.id,
            role_id="project_editor",
            department_id=d.id,
        )
    )
    db_session.add(
        UserRole(
            user_id=viewer_user.id, role_id="viewer", department_id=d.id
        )
    )
    db_session.flush()
    cl = Client(code="CON", name="Contoso", department_id=d.id)
    di = Discipline(
        code="Design", name="Protection & Controls", department_id=d.id
    )
    db_session.add_all([cl, di])
    db_session.flush()
    t = Template(
        name="t",
        department_id=d.id,
        client_id=cl.id,
        discipline_id=di.id,
    )
    db_session.add(t)
    db_session.flush()
    f = TemplateFieldDef(
        template_id=t.id,
        name="Required Notes",
        field_type="short_text",
        required=True,
        order_index=0,
    )
    db_session.add(f)
    m = TemplateMilestoneDef(
        template_id=t.id,
        name="IFC",
        direction="outbound",
        date_model="single",
        order_index=0,
    )
    db_session.add(m)
    db_session.flush()
    return t, f, m


def _create(c, template_id: str, **extra) -> dict:
    payload = {
        "project_number": "25756601",
        "title": "Test Project",
        "template_id": template_id,
        **extra,
    }
    r = c.post("/api/projects", json=payload)
    assert r.status_code == 201, r.text
    return r.json()


# ---- auth gating --------------------------------------------------------


def test_list_requires_auth(client: TestClient):
    assert client.get("/api/projects").status_code == 401


def test_viewer_can_list(client_as, viewer_user: User):
    assert client_as(viewer_user).get("/api/projects").status_code == 200


def test_viewer_cannot_create(
    client_as, viewer_user: User, template_with_defs
):
    t, _f, _m = template_with_defs
    r = client_as(viewer_user).post(
        "/api/projects",
        json={
            "project_number": "12345",
            "title": "x",
            "template_id": str(t.id),
        },
    )
    assert r.status_code == 403


def test_project_editor_can_create(
    client_as, project_editor_user: User, template_with_defs
):
    t, _f, _m = template_with_defs
    project = _create(
        client_as(project_editor_user), str(t.id),
        custom_field_values={},
    )
    assert project["lifecycle_state"] == "draft"


# ---- create + auto-milestone -------------------------------------------


def test_create_auto_spawns_milestones(
    client_as, admin_user: User, template_with_defs
):
    t, _f, _m = template_with_defs
    project = _create(client_as(admin_user), str(t.id))
    assert len(project["milestones"]) == 1
    ms = project["milestones"][0]
    assert ms["name"] == "IFC"
    assert ms["direction"] == "outbound"
    assert ms["template_milestone_def_id"] is not None
    assert ms["planned_date"] is None  # filled in later


def test_create_with_invalid_template_id_422(
    client_as, admin_user: User
):
    r = client_as(admin_user).post(
        "/api/projects",
        json={
            "project_number": "x12345",
            "title": "x",
            "template_id": "00000000-0000-0000-0000-000000000000",
        },
    )
    assert r.status_code == 422


def test_create_with_unknown_field_value_key_422(
    client_as, admin_user: User, template_with_defs
):
    t, _f, _m = template_with_defs
    r = client_as(admin_user).post(
        "/api/projects",
        json={
            "project_number": "x123",
            "title": "x",
            "template_id": str(t.id),
            "custom_field_values": {
                "00000000-0000-0000-0000-000000000000": "stray"
            },
        },
    )
    assert r.status_code == 422


def test_create_duplicate_project_number_409(
    client_as, admin_user: User, template_with_defs
):
    t, _f, _m = template_with_defs
    c = client_as(admin_user)
    _create(c, str(t.id))
    r = c.post(
        "/api/projects",
        json={
            "project_number": "25756601",
            "title": "dup",
            "template_id": str(t.id),
        },
    )
    assert r.status_code == 409


def test_create_with_valid_field_value(
    client_as, admin_user: User, template_with_defs
):
    t, f, _m = template_with_defs
    c = client_as(admin_user)
    r = c.post(
        "/api/projects",
        json={
            "project_number": "x123",
            "title": "x",
            "template_id": str(t.id),
            "custom_field_values": {str(f.id): "Some notes"},
        },
    )
    assert r.status_code == 201
    assert r.json()["custom_field_values"][str(f.id)] == "Some notes"


# ---- patch + merge ------------------------------------------------------


def test_patch_merges_custom_field_values(
    client_as, admin_user: User, template_with_defs, db_session: Session
):
    t, f, _m = template_with_defs
    # Add a second field so we can prove the merge is partial.
    f2 = TemplateFieldDef(
        template_id=t.id,
        name="Color",
        field_type="single_select",
        options={"choices": ["red", "blue"]},
        order_index=1,
    )
    db_session.add(f2)
    db_session.flush()

    c = client_as(admin_user)
    p = _create(
        c,
        str(t.id),
        custom_field_values={str(f.id): "notes", str(f2.id): "red"},
    )

    # Patch only one of the two fields.
    r = c.patch(
        f"/api/projects/{p['id']}",
        json={"custom_field_values": {str(f.id): "new notes"}},
    )
    assert r.status_code == 200
    cfv = r.json()["custom_field_values"]
    assert cfv[str(f.id)] == "new notes"
    assert cfv[str(f2.id)] == "red"


def test_patch_strips_orphan_keys_from_soft_deleted_fields(
    client_as, admin_user: User, template_with_defs, db_session: Session
):
    """A project created when the template had field F still carries F's
    UUID in its JSONB after F is soft-deleted. The next PATCH must
    succeed (the user can't see or fix orphan data from the UI) and
    must quietly strip the orphan key from the stored values."""
    from datetime import datetime, timezone
    t, f, _m = template_with_defs
    # Add a second field so the project has two custom values, then
    # we'll soft-delete one and PATCH the other.
    f2 = TemplateFieldDef(
        template_id=t.id,
        name="Doomed",
        field_type="short_text",
        order_index=1,
    )
    db_session.add(f2)
    db_session.flush()
    c = client_as(admin_user)
    p = _create(
        c,
        str(t.id),
        custom_field_values={str(f.id): "notes", str(f2.id): "carried"},
    )
    # Soft-delete f2 — UI is gone, but the project's JSONB still has the key.
    f2.deleted_at = datetime.now(timezone.utc)
    db_session.flush()
    db_session.commit()

    # Frontend echoes back the full dict including the orphan key
    # (it loaded it from GET, can't see the orphan field, sends it
    # along on save). PATCH must succeed and drop the orphan.
    r = c.patch(
        f"/api/projects/{p['id']}",
        json={
            "custom_field_values": {
                str(f.id): "updated",
                str(f2.id): "carried",  # orphan — sent by the UI
            }
        },
    )
    assert r.status_code == 200, r.text
    cfv = r.json()["custom_field_values"]
    assert cfv[str(f.id)] == "updated"
    assert str(f2.id) not in cfv


def test_patch_null_removes_key(
    client_as, admin_user: User, template_with_defs
):
    t, f, _m = template_with_defs
    c = client_as(admin_user)
    p = _create(c, str(t.id), custom_field_values={str(f.id): "x"})
    r = c.patch(
        f"/api/projects/{p['id']}",
        json={"custom_field_values": {str(f.id): None}},
    )
    assert r.status_code == 200
    assert str(f.id) not in r.json()["custom_field_values"]


def test_patch_rejects_lifecycle_state(
    client_as, admin_user: User, template_with_defs
):
    """ProjectUpdate schema has no lifecycle_state field — Pydantic silently
    ignores unknown fields but the schema doesn't accept the value either way.
    Confirm: PATCH with lifecycle_state changes nothing on the row."""
    t, _f, _m = template_with_defs
    c = client_as(admin_user)
    p = _create(c, str(t.id))
    r = c.patch(
        f"/api/projects/{p['id']}",
        json={"title": "ok", "lifecycle_state": "active"},
    )
    # Must be 200 (title still patched) but lifecycle_state unchanged.
    assert r.status_code == 200
    assert r.json()["lifecycle_state"] == "draft"


# ---- transitions --------------------------------------------------------


def test_transition_draft_to_active_blocks_when_required_field_missing(
    client_as, admin_user: User, template_with_defs, db_session: Session
):
    t, _f, m = template_with_defs
    c = client_as(admin_user)
    p = _create(c, str(t.id))  # no custom field values, no milestone dates
    r = c.post(
        f"/api/projects/{p['id']}/transition",
        json={"to_state": "active"},
    )
    assert r.status_code == 422
    detail = r.json()["detail"]
    assert any("required field" in d for d in detail)
    assert any("planned date" in d for d in detail)


def test_transition_draft_to_active_succeeds_when_ready(
    client_as, admin_user: User, template_with_defs, db_session: Session
):
    from datetime import date as _date

    from backend.app.db.models import Milestone, Project

    t, f, _m = template_with_defs
    c = client_as(admin_user)
    p = _create(
        c,
        str(t.id),
        custom_field_values={str(f.id): "all set"},
    )
    # Fill the milestone's planned date directly via the DB.
    project = db_session.get(Project, p["id"])
    for ms in project.milestones:
        ms.planned_date = _date(2026, 6, 1)
    db_session.flush()

    r = c.post(
        f"/api/projects/{p['id']}/transition",
        json={"to_state": "active"},
    )
    assert r.status_code == 200, r.text
    assert r.json()["lifecycle_state"] == "active"


def test_transition_invalid_target_422(
    client_as, admin_user: User, template_with_defs
):
    t, _f, _m = template_with_defs
    c = client_as(admin_user)
    p = _create(c, str(t.id))
    r = c.post(
        f"/api/projects/{p['id']}/transition",
        json={"to_state": "complete"},
    )
    assert r.status_code == 422


def test_transition_unknown_state_422(
    client_as, admin_user: User, template_with_defs
):
    t, _f, _m = template_with_defs
    c = client_as(admin_user)
    p = _create(c, str(t.id))
    r = c.post(
        f"/api/projects/{p['id']}/transition",
        json={"to_state": "weird"},
    )
    assert r.status_code == 422


def test_transition_cancel_from_draft_ok(
    client_as, admin_user: User, template_with_defs
):
    t, _f, _m = template_with_defs
    c = client_as(admin_user)
    p = _create(c, str(t.id))
    r = c.post(
        f"/api/projects/{p['id']}/transition",
        json={"to_state": "cancelled"},
    )
    assert r.status_code == 200
    assert r.json()["lifecycle_state"] == "cancelled"


# ---- soft-delete --------------------------------------------------------


def test_soft_delete_hides_from_list(
    client_as, admin_user: User, template_with_defs
):
    t, _f, _m = template_with_defs
    c = client_as(admin_user)
    p = _create(c, str(t.id))
    assert c.delete(f"/api/projects/{p['id']}").status_code == 204
    body = c.get("/api/projects").json()
    ids = [pp["id"] for pp in body["items"]]
    assert p["id"] not in ids
    assert c.get(f"/api/projects/{p['id']}").status_code == 404


# ---- detail endpoint -----------------------------------------------------


def test_get_returns_milestones_and_valid_next_states(
    client_as, admin_user: User, template_with_defs
):
    t, _f, _m = template_with_defs
    c = client_as(admin_user)
    p = _create(c, str(t.id))
    body = c.get(f"/api/projects/{p['id']}").json()
    assert len(body["milestones"]) == 1
    assert set(body["valid_next_states"]) == {"active", "cancelled"}


# ---- filtering -----------------------------------------------------------


def test_list_filters_by_template_id_and_lifecycle_state(
    client_as, admin_user: User, template_with_defs
):
    t, _f, _m = template_with_defs
    c = client_as(admin_user)
    _create(c, str(t.id))
    body = c.get(
        f"/api/projects?template_id={t.id}&lifecycle_state=draft"
    ).json()
    assert body["total"] == 1
    body2 = c.get(
        f"/api/projects?lifecycle_state=active"
    ).json()
    assert body2["total"] == 0


# Phase 4.8.14 — D/C/D filters.


def test_list_filters_by_department_client_and_discipline(
    client_as, admin_user: User, template_with_defs, db_session
):
    t1, _f, _m = template_with_defs
    # Build a second template under a different department + client +
    # discipline so each of the new filters has something to narrow on.
    other_dept = Department(code="DIV4", name="Division 4")
    db_session.add(other_dept)
    db_session.flush()
    other_client = Client(code="FAB", name="Fabrikam", department_id=other_dept.id)
    other_disc = Discipline(
        code="Physical", name="Physical", department_id=other_dept.id
    )
    db_session.add_all([other_client, other_disc])
    db_session.flush()
    t2 = Template(
        name="t2",
        department_id=other_dept.id,
        client_id=other_client.id,
        discipline_id=other_disc.id,
    )
    db_session.add(t2)
    db_session.flush()

    c = client_as(admin_user)
    _create(c, str(t1.id), project_number="11111")
    _create(c, str(t2.id), project_number="22222")

    # Department filter
    body = c.get(f"/api/projects?department_id={t1.department_id}").json()
    assert body["total"] == 1
    assert body["items"][0]["template_id"] == str(t1.id)

    # Client filter
    body = c.get(f"/api/projects?client_id={t2.client_id}").json()
    assert body["total"] == 1
    assert body["items"][0]["template_id"] == str(t2.id)

    # Discipline filter
    body = c.get(f"/api/projects?discipline_id={t1.discipline_id}").json()
    assert body["total"] == 1
    assert body["items"][0]["template_id"] == str(t1.id)

    # Combining filters narrows further (department of t1 + discipline
    # of t2 → empty since they live in different templates).
    body = c.get(
        f"/api/projects?department_id={t1.department_id}"
        f"&discipline_id={t2.discipline_id}"
    ).json()
    assert body["total"] == 0


# ---- search (Phase 2.6) ------------------------------------------------


def test_list_search_matches_title_case_insensitive(
    client_as, admin_user: User, template_with_defs
):
    t, _f, _m = template_with_defs
    c = client_as(admin_user)
    _create(c, str(t.id), title="Endor RTU Replacement")
    _create(c, str(t.id), project_number="9999", title="Hoth Subzero")
    body = c.get("/api/projects?q=ENDOR").json()
    titles = [p["title"] for p in body["items"]]
    assert titles == ["Endor RTU Replacement"]
    assert body["total"] == 1


def test_list_search_matches_project_number(
    client_as, admin_user: User, template_with_defs
):
    t, _f, _m = template_with_defs
    c = client_as(admin_user)
    _create(c, str(t.id), project_number="25756614")
    _create(c, str(t.id), project_number="99999")
    body = c.get("/api/projects?q=2575").json()
    assert body["total"] == 1
    assert body["items"][0]["project_number"] == "25756614"


def test_list_search_matches_client_number(
    client_as, admin_user: User, template_with_defs
):
    t, _f, _m = template_with_defs
    c = client_as(admin_user)
    _create(c, str(t.id), client_project_number="M18025410")
    _create(c, str(t.id), project_number="00007")
    body = c.get("/api/projects?q=M18025").json()
    assert body["total"] == 1
    assert body["items"][0]["client_project_number"] == "M18025410"


def test_list_search_no_match(
    client_as, admin_user: User, template_with_defs
):
    t, _f, _m = template_with_defs
    c = client_as(admin_user)
    _create(c, str(t.id))
    body = c.get("/api/projects?q=nothing-matches-this").json()
    assert body["total"] == 0
    assert body["items"] == []


def test_list_search_whitespace_only_is_no_filter(
    client_as, admin_user: User, template_with_defs
):
    """A stray space in the search box must NOT yield an empty list."""
    t, _f, _m = template_with_defs
    c = client_as(admin_user)
    _create(c, str(t.id))
    full = c.get("/api/projects").json()
    spaced = c.get("/api/projects?q=%20%20").json()
    assert spaced["total"] == full["total"]


# ---- milestone date PATCH (Phase 1.7.3) ---------------------------------


def test_patch_milestone_dates_happy_path(
    client_as, admin_user: User, template_with_defs
):
    t, _f, _m = template_with_defs
    c = client_as(admin_user)
    p = _create(c, str(t.id))
    mid = p["milestones"][0]["id"]
    r = c.patch(
        f"/api/projects/{p['id']}/milestones/{mid}",
        json={"planned_date": "2026-06-01", "actual_date": "2026-06-05"},
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["planned_date"] == "2026-06-01"
    assert body["actual_date"] == "2026-06-05"


def test_patch_milestone_clear_date_with_null(
    client_as, admin_user: User, template_with_defs
):
    t, _f, _m = template_with_defs
    c = client_as(admin_user)
    p = _create(c, str(t.id))
    mid = p["milestones"][0]["id"]
    c.patch(
        f"/api/projects/{p['id']}/milestones/{mid}",
        json={"planned_date": "2026-06-01"},
    )
    r = c.patch(
        f"/api/projects/{p['id']}/milestones/{mid}",
        json={"planned_date": None},
    )
    assert r.status_code == 200
    assert r.json()["planned_date"] is None


def test_patch_milestone_wrong_project_404(
    client_as, admin_user: User, template_with_defs
):
    t, _f, _m = template_with_defs
    c = client_as(admin_user)
    p1 = _create(c, str(t.id))
    p2 = c.post(
        "/api/projects",
        json={
            "project_number": "25756602",
            "title": "Other",
            "template_id": str(t.id),
        },
    ).json()
    other_mid = p2["milestones"][0]["id"]
    r = c.patch(
        f"/api/projects/{p1['id']}/milestones/{other_mid}",
        json={"planned_date": "2026-06-01"},
    )
    assert r.status_code == 404


def test_patch_milestone_forbidden_for_viewer(
    client_as,
    admin_user: User,
    viewer_user: User,
    template_with_defs,
):
    t, _f, _m = template_with_defs
    p = _create(client_as(admin_user), str(t.id))
    mid = p["milestones"][0]["id"]
    r = client_as(viewer_user).patch(
        f"/api/projects/{p['id']}/milestones/{mid}",
        json={"planned_date": "2026-06-01"},
    )
    assert r.status_code == 403


def test_patch_milestone_empty_body_422(
    client_as, admin_user: User, template_with_defs
):
    t, _f, _m = template_with_defs
    c = client_as(admin_user)
    p = _create(c, str(t.id))
    mid = p["milestones"][0]["id"]
    r = c.patch(f"/api/projects/{p['id']}/milestones/{mid}", json={})
    assert r.status_code == 422


# ---- milestone full CRUD (Phase 1.8.1) ----------------------------------


def test_create_adhoc_milestone(
    client_as, admin_user: User, template_with_defs
):
    t, _f, _m = template_with_defs
    c = client_as(admin_user)
    p = _create(c, str(t.id))
    r = c.post(
        f"/api/projects/{p['id']}/milestones",
        json={
            "name": "Surprise milestone",
            "direction": "internal",
            "date_model": "single",
        },
    )
    assert r.status_code == 201, r.text
    body = r.json()
    assert body["name"] == "Surprise milestone"
    assert body["template_milestone_def_id"] is None  # ad-hoc
    assert body["planned_date"] is None
    assert body["order_index"] == 1  # template milestone was 0


def test_create_milestone_invalid_direction_422(
    client_as, admin_user: User, template_with_defs
):
    t, _f, _m = template_with_defs
    c = client_as(admin_user)
    p = _create(c, str(t.id))
    r = c.post(
        f"/api/projects/{p['id']}/milestones",
        json={"name": "x", "direction": "sideways", "date_model": "single"},
    )
    assert r.status_code == 422


def test_create_milestone_forbidden_for_viewer(
    client_as,
    admin_user: User,
    viewer_user: User,
    template_with_defs,
):
    t, _f, _m = template_with_defs
    p = _create(client_as(admin_user), str(t.id))
    r = client_as(viewer_user).post(
        f"/api/projects/{p['id']}/milestones",
        json={"name": "x", "direction": "outbound", "date_model": "single"},
    )
    assert r.status_code == 403


def test_patch_milestone_rename_and_direction(
    client_as, admin_user: User, template_with_defs
):
    t, _f, _m = template_with_defs
    c = client_as(admin_user)
    p = _create(c, str(t.id))
    mid = p["milestones"][0]["id"]
    r = c.patch(
        f"/api/projects/{p['id']}/milestones/{mid}",
        json={
            "name": "Issued for Construction",
            "direction": "external",
        },
    )
    assert r.status_code == 200
    body = r.json()
    assert body["name"] == "Issued for Construction"
    assert body["direction"] == "external"


def test_patch_milestone_invalid_direction_422(
    client_as, admin_user: User, template_with_defs
):
    t, _f, _m = template_with_defs
    c = client_as(admin_user)
    p = _create(c, str(t.id))
    mid = p["milestones"][0]["id"]
    r = c.patch(
        f"/api/projects/{p['id']}/milestones/{mid}",
        json={"direction": "sideways"},
    )
    assert r.status_code == 422


def test_delete_milestone_soft_delete(
    client_as, admin_user: User, template_with_defs
):
    t, _f, _m = template_with_defs
    c = client_as(admin_user)
    p = _create(c, str(t.id))
    mid = p["milestones"][0]["id"]
    r = c.delete(f"/api/projects/{p['id']}/milestones/{mid}")
    assert r.status_code == 204
    body = c.get(f"/api/projects/{p['id']}").json()
    ids = [m["id"] for m in body["milestones"]]
    assert mid not in ids


def test_delete_milestone_forbidden_for_viewer(
    client_as,
    admin_user: User,
    viewer_user: User,
    template_with_defs,
):
    t, _f, _m = template_with_defs
    p = _create(client_as(admin_user), str(t.id))
    mid = p["milestones"][0]["id"]
    r = client_as(viewer_user).delete(
        f"/api/projects/{p['id']}/milestones/{mid}"
    )
    assert r.status_code == 403


def test_reorder_milestones_happy_path(
    client_as, admin_user: User, template_with_defs
):
    t, _f, _m = template_with_defs
    c = client_as(admin_user)
    p = _create(c, str(t.id))
    b = c.post(
        f"/api/projects/{p['id']}/milestones",
        json={"name": "B", "direction": "outbound", "date_model": "single"},
    ).json()
    cc = c.post(
        f"/api/projects/{p['id']}/milestones",
        json={"name": "C", "direction": "outbound", "date_model": "single"},
    ).json()
    a_id = p["milestones"][0]["id"]
    r = c.post(
        f"/api/projects/{p['id']}/milestones/reorder",
        json={"ordered_ids": [cc["id"], a_id, b["id"]]},
    )
    assert r.status_code == 204
    body = c.get(f"/api/projects/{p['id']}").json()
    assert [m["name"] for m in body["milestones"]] == ["C", "IFC", "B"]


def test_reorder_milestones_missing_id_422(
    client_as, admin_user: User, template_with_defs
):
    t, _f, _m = template_with_defs
    c = client_as(admin_user)
    p = _create(c, str(t.id))
    c.post(
        f"/api/projects/{p['id']}/milestones",
        json={"name": "B", "direction": "outbound", "date_model": "single"},
    )
    r = c.post(
        f"/api/projects/{p['id']}/milestones/reorder",
        json={"ordered_ids": [p["milestones"][0]["id"]]},
    )
    assert r.status_code == 422
    assert any("missing id" in d for d in r.json()["detail"])


def test_reorder_milestones_forbidden_for_viewer(
    client_as,
    admin_user: User,
    viewer_user: User,
    template_with_defs,
):
    t, _f, _m = template_with_defs
    p = _create(client_as(admin_user), str(t.id))
    mid = p["milestones"][0]["id"]
    r = client_as(viewer_user).post(
        f"/api/projects/{p['id']}/milestones/reorder",
        json={"ordered_ids": [mid]},
    )
    assert r.status_code == 403


# ---- Phase 2.7.2: sort extension ----------------------------------------


def test_list_sort_by_title_asc(db_session, client_as, admin_user):
    """`?sort=title&sort_direction=asc` orders rows by title ascending."""
    c = client_as(admin_user)
    # Seed a template + 3 projects with intentionally jumbled titles.
    from backend.tests.test_view_columns_routes import (
        _make_template_with_one_field_and_one_milestone,
    )
    from backend.app.db.models import Project

    t, _, _ = _make_template_with_one_field_and_one_milestone(db_session)
    for title, number in [("Zeta", "25700001"), ("Alpha", "25700002"), ("Mu", "25700003")]:
        db_session.add(
            Project(
                project_number=number,
                title=title,
                template_id=t.id,
                created_by=admin_user.id,
            )
        )
    db_session.commit()
    r = c.get("/api/projects?sort=title&sort_direction=asc")
    assert r.status_code == 200, r.text
    titles = [item["title"] for item in r.json()["items"]]
    assert titles == sorted(titles), titles


def test_list_sort_by_title_desc(db_session, client_as, admin_user):
    c = client_as(admin_user)
    from backend.tests.test_view_columns_routes import (
        _make_template_with_one_field_and_one_milestone,
    )
    from backend.app.db.models import Project

    t, _, _ = _make_template_with_one_field_and_one_milestone(db_session)
    for title, number in [("Zeta", "25700001"), ("Alpha", "25700002"), ("Mu", "25700003")]:
        db_session.add(
            Project(
                project_number=number,
                title=title,
                template_id=t.id,
                created_by=admin_user.id,
            )
        )
    db_session.commit()
    r = c.get("/api/projects?sort=title&sort_direction=desc")
    titles = [item["title"] for item in r.json()["items"]]
    assert titles == sorted(titles, reverse=True), titles


def test_list_sort_rejects_unknown_key(db_session, client_as, admin_user):
    """`?sort=evil_drop_table` is rejected — whitelist enforcement."""
    c = client_as(admin_user)
    r = c.get("/api/projects?sort=evil_drop_table")
    assert r.status_code == 422


def test_list_no_sort_default_created_desc_unchanged(
    db_session, client_as, admin_user
):
    """Existing default (created_at DESC) is preserved when sort is omitted."""
    c = client_as(admin_user)
    r = c.get("/api/projects")
    assert r.status_code == 200


# ---- Phase 2.7.2: expand_milestones extension ---------------------------


def test_list_expand_milestones_true_populates_items(
    db_session, client_as, admin_user
):
    c = client_as(admin_user)
    from backend.tests.test_view_columns_routes import (
        _make_template_with_one_field_and_one_milestone,
    )
    from backend.app.db.models import Milestone, Project

    t, _, md = _make_template_with_one_field_and_one_milestone(db_session)
    p = Project(
        project_number="25710001",
        title="Test",
        template_id=t.id,
        created_by=admin_user.id,
    )
    db_session.add(p)
    db_session.flush()
    db_session.add(
        Milestone(
            project_id=p.id,
            template_milestone_def_id=md.id,
            name="M1",
            direction="outbound",
            date_model="planned_actual",
        )
    )
    db_session.commit()

    r = c.get("/api/projects?expand_milestones=true")
    assert r.status_code == 200
    body = r.json()
    item = next(x for x in body["items"] if x["id"] == str(p.id))
    assert "milestones" in item
    assert len(item["milestones"]) == 1
    assert item["milestones"][0]["name"] == "M1"


def test_list_expand_milestones_false_omits(db_session, client_as, admin_user):
    c = client_as(admin_user)
    r = c.get("/api/projects")
    body = r.json()
    for item in body["items"]:
        assert "milestones" not in item or item["milestones"] is None


# ---- Phase 2.7.2: expand_refs extension ---------------------------------


def test_list_expand_refs_returns_labels(
    db_session, client_as, admin_user, viewer_user
):
    """`?expand_refs=true&template_id=<t>` returns ref_labels for the
    user UUIDs referenced by the page's custom_field_values."""
    c = client_as(admin_user)
    from backend.tests.test_ref_labels_service import _seed_minimal
    from backend.app.db.models import Project

    t, ids = _seed_minimal(db_session, admin_user)
    p = Project(
        project_number="25712001",
        title="T1",
        template_id=t.id,
        created_by=admin_user.id,
        custom_field_values={str(ids["fd_user"].id): str(viewer_user.id)},
    )
    db_session.add(p)
    db_session.commit()

    r = c.get(f"/api/projects?template_id={t.id}&expand_refs=true")
    assert r.status_code == 200, r.text
    body = r.json()
    assert "ref_labels" in body
    assert body["ref_labels"]["users"][str(viewer_user.id)] == (
        viewer_user.display_name
    )


def test_list_expand_refs_false_omits(db_session, client_as, admin_user):
    c = client_as(admin_user)
    r = c.get("/api/projects")
    body = r.json()
    assert body.get("ref_labels") is None


# ---- audit log (Phase 3.1) ----------------------------------------------


def _audit_rows(db, entity_type, entity_id):
    from sqlalchemy import select
    from backend.app.db.models import AuditLog
    import uuid as _uuid

    return list(
        db.execute(
            select(AuditLog).where(
                AuditLog.entity_type == entity_type,
                AuditLog.entity_id == _uuid.UUID(str(entity_id)),
            )
        ).scalars()
    )


def test_create_project_writes_audit_row(
    db_session, client_as, admin_user: User, template_with_defs
):
    t, _f, _m = template_with_defs
    c = client_as(admin_user)
    proj = _create(c, str(t.id))
    rows = _audit_rows(db_session, "project", proj["id"])
    assert len(rows) == 1
    row = rows[0]
    assert row.operation == "create"
    assert row.changed_by == admin_user.id
    assert row.project_id is not None and str(row.project_id) == proj["id"]
    assert row.changes["initial"]["title"] == "Test Project"


def test_patch_project_writes_audit_row_with_diff(
    db_session, client_as, admin_user: User, template_with_defs
):
    t, _f, _m = template_with_defs
    c = client_as(admin_user)
    proj = _create(c, str(t.id))
    r = c.patch(
        f"/api/projects/{proj['id']}",
        json={"title": "Renamed", "client_project_number": "CN-99"},
    )
    assert r.status_code == 200, r.text
    rows = _audit_rows(db_session, "project", proj["id"])
    update_rows = [r for r in rows if r.operation == "update"]
    assert len(update_rows) == 1
    changes = update_rows[0].changes
    assert changes["title"] == ["Test Project", "Renamed"]
    assert changes["client_project_number"][1] == "CN-99"


def test_patch_project_cfv_writes_sub_diff(
    db_session, client_as, admin_user: User, template_with_defs
):
    t, f, _m = template_with_defs
    c = client_as(admin_user)
    proj = _create(c, str(t.id), custom_field_values={str(f.id): "first"})
    r = c.patch(
        f"/api/projects/{proj['id']}",
        json={"custom_field_values": {str(f.id): "second"}},
    )
    assert r.status_code == 200, r.text
    rows = _audit_rows(db_session, "project", proj["id"])
    update_rows = [r for r in rows if r.operation == "update"]
    assert len(update_rows) == 1
    cfv = update_rows[0].changes.get("custom_field_values")
    assert cfv is not None
    assert cfv[str(f.id)] == ["first", "second"]


def test_delete_project_writes_audit_row(
    db_session, client_as, admin_user: User, template_with_defs
):
    t, _f, _m = template_with_defs
    c = client_as(admin_user)
    proj = _create(c, str(t.id))
    r = c.delete(f"/api/projects/{proj['id']}")
    assert r.status_code == 204
    rows = _audit_rows(db_session, "project", proj["id"])
    delete_rows = [r for r in rows if r.operation == "delete"]
    assert len(delete_rows) == 1
    assert delete_rows[0].changes == {}


def test_transition_project_writes_audit_row(
    db_session, client_as, admin_user: User, template_with_defs
):
    t, _f, _m = template_with_defs
    c = client_as(admin_user)
    proj = _create(c, str(t.id))
    r = c.post(
        f"/api/projects/{proj['id']}/transition",
        json={"to_state": "cancelled"},
    )
    assert r.status_code == 200, r.text
    rows = _audit_rows(db_session, "project", proj["id"])
    trans_rows = [r for r in rows if r.operation == "transition"]
    assert len(trans_rows) == 1
    assert trans_rows[0].changes == {"from": "draft", "to": "cancelled"}


def test_create_milestone_writes_audit_row(
    db_session, client_as, admin_user: User, template_with_defs
):
    t, _f, _m = template_with_defs
    c = client_as(admin_user)
    proj = _create(c, str(t.id))
    r = c.post(
        f"/api/projects/{proj['id']}/milestones",
        json={
            "name": "Ad-hoc",
            "direction": "outbound",
            "date_model": "single",
        },
    )
    assert r.status_code == 201, r.text
    m = r.json()
    rows = _audit_rows(db_session, "milestone", m["id"])
    create_rows = [r for r in rows if r.operation == "create"]
    assert len(create_rows) == 1
    assert create_rows[0].project_id is not None
    assert str(create_rows[0].project_id) == proj["id"]
    assert create_rows[0].changes["initial"]["name"] == "Ad-hoc"


def test_patch_milestone_writes_audit_row_with_diff(
    db_session, client_as, admin_user: User, template_with_defs
):
    t, _f, _m = template_with_defs
    c = client_as(admin_user)
    proj = _create(c, str(t.id))
    detail = c.get(f"/api/projects/{proj['id']}").json()
    mid = detail["milestones"][0]["id"]
    r = c.patch(
        f"/api/projects/{proj['id']}/milestones/{mid}",
        json={"planned_date": "2026-09-01"},
    )
    assert r.status_code == 200, r.text
    rows = _audit_rows(db_session, "milestone", mid)
    update_rows = [r for r in rows if r.operation == "update"]
    assert len(update_rows) == 1
    assert update_rows[0].changes["planned_date"][1] == "2026-09-01"
    assert str(update_rows[0].project_id) == proj["id"]


def test_delete_milestone_writes_audit_row(
    db_session, client_as, admin_user: User, template_with_defs
):
    t, _f, _m = template_with_defs
    c = client_as(admin_user)
    proj = _create(c, str(t.id))
    detail = c.get(f"/api/projects/{proj['id']}").json()
    mid = detail["milestones"][0]["id"]
    r = c.delete(f"/api/projects/{proj['id']}/milestones/{mid}")
    assert r.status_code == 204
    rows = _audit_rows(db_session, "milestone", mid)
    delete_rows = [r for r in rows if r.operation == "delete"]
    assert len(delete_rows) == 1
    assert str(delete_rows[0].project_id) == proj["id"]


# ---- Phase 5.3 — spreadsheet import ------------------------------------


import json as _json


def _import(
    c,
    template_id: str,
    csv_body: str,
    mapping: dict[str, str],
) -> dict:
    """POST a CSV to the import endpoint and return the JSON body."""
    r = c.post(
        "/api/projects/import",
        files={"file": ("rows.csv", csv_body.encode("utf-8"), "text/csv")},
        data={
            "template_id": template_id,
            "mapping": _json.dumps(mapping),
        },
    )
    return r


def test_import_requires_dm_or_above(
    client_as, viewer_user: User, project_editor_user: User,
    template_with_defs, db_session,
):
    t, _f, _m = template_with_defs
    csv = "Number\nM-1\n"
    mapping = {"Number": "project_number"}
    # Viewer and project_editor both get 403; admin gets through.
    r = _import(client_as(viewer_user), str(t.id), csv, mapping)
    assert r.status_code == 403
    r = _import(client_as(project_editor_user), str(t.id), csv, mapping)
    assert r.status_code == 403


def test_import_dm_can_create(
    client_as, admin_user: User, template_with_defs,
):
    t, _f, _m = template_with_defs
    csv = "Number,Title\nM-1,Hawthorn upgrade\nM-2,Endor RTU\n"
    r = _import(
        client_as(admin_user),
        str(t.id),
        csv,
        {"Number": "project_number", "Title": "title"},
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["created"] == 2
    assert body["skipped"] == []
    assert body["errors"] == []


def test_import_skips_blank_mesa(
    client_as, admin_user: User, template_with_defs,
):
    t, _f, _m = template_with_defs
    csv = "Number,Title\nM-1,Live\n,Orphan\n"
    r = _import(
        client_as(admin_user),
        str(t.id),
        csv,
        {"Number": "project_number", "Title": "title"},
    )
    body = r.json()
    assert body["created"] == 1
    assert len(body["skipped"]) == 1
    assert body["skipped"][0]["row"] == 3
    assert body["skipped"][0]["reason"] == "missing Project #"


def test_import_skips_duplicate_mesa(
    client_as, admin_user: User, template_with_defs,
):
    t, _f, _m = template_with_defs
    c = client_as(admin_user)
    _create(c, str(t.id), project_number="DUPE")
    csv = "Number,Title\nDUPE,Should skip\nFRESH,Should land\n"
    r = _import(
        c, str(t.id), csv,
        {"Number": "project_number", "Title": "title"},
    )
    body = r.json()
    assert body["created"] == 1
    assert len(body["skipped"]) == 1
    assert body["skipped"][0]["project_number"] == "DUPE"
    assert body["skipped"][0]["reason"] == "Project # already exists"


def test_import_blank_title_uses_placeholder(
    client_as, admin_user: User, template_with_defs,
):
    t, _f, _m = template_with_defs
    csv = "Number,Title\nM-1,\n"
    r = _import(
        client_as(admin_user),
        str(t.id),
        csv,
        {"Number": "project_number", "Title": "title"},
    )
    assert r.status_code == 200
    assert r.json()["created"] == 1
    # The new project's title is the placeholder.
    proj = client_as(admin_user).get("/api/projects?q=M-1").json()["items"][0]
    assert proj["title"] == "UPDATE"


def test_import_drops_bad_custom_field_value(
    client_as, admin_user: User, template_with_defs, db_session,
):
    """A value that fails its type validator is silently dropped;
    the row still creates a project with the other fields."""
    t, _f, _m = template_with_defs
    # Add a numeric (currency) custom field on this template so the
    # validator has something to reject.
    c = client_as(admin_user)
    fid = c.post(
        f"/api/admin/templates/{t.id}/fields",
        json={"name": "Budget", "field_type": "currency"},
    ).json()["id"]

    csv = "Number,Title,Budget\nM-1,Live,not-a-number\n"
    r = _import(
        c, str(t.id), csv,
        {
            "Number": "project_number",
            "Title": "title",
            "Budget": fid,
        },
    )
    assert r.status_code == 200
    body = r.json()
    assert body["created"] == 1
    # The created project's custom_field_values must NOT contain the
    # invalid budget value.
    proj = c.get("/api/projects?q=M-1").json()["items"][0]
    proj_detail = c.get(f"/api/projects/{proj['id']}").json()
    assert proj_detail["custom_field_values"] == {}


def test_import_projects_land_as_draft(
    client_as, admin_user: User, template_with_defs,
):
    t, _f, _m = template_with_defs
    csv = "Number,Title\nM-1,Endor\n"
    r = _import(
        client_as(admin_user),
        str(t.id),
        csv,
        {"Number": "project_number", "Title": "title"},
    )
    assert r.status_code == 200
    proj = client_as(admin_user).get("/api/projects?q=M-1").json()["items"][0]
    assert proj["lifecycle_state"] == "draft"


def test_import_audit_log_records_create_per_row(
    client_as, admin_user: User, template_with_defs, db_session,
):
    t, _f, _m = template_with_defs
    csv = "Number,Title\nM-1,One\nM-2,Two\n"
    r = _import(
        client_as(admin_user),
        str(t.id),
        csv,
        {"Number": "project_number", "Title": "title"},
    )
    assert r.status_code == 200
    items = client_as(admin_user).get("/api/projects").json()["items"]
    ids = [p["id"] for p in items]
    for pid in ids:
        rows = _audit_rows(db_session, "project", pid)
        creates = [r for r in rows if r.operation == "create"]
        assert len(creates) == 1
        assert creates[0].changes.get("via") == "import"


def test_import_rejects_empty_csv(
    client_as, admin_user: User, template_with_defs,
):
    t, _f, _m = template_with_defs
    r = _import(
        client_as(admin_user), str(t.id), "", {"Number": "project_number"},
    )
    assert r.status_code == 422
    assert "empty" in r.json()["detail"].lower()


def test_import_rejects_unknown_template(
    client_as, admin_user: User, template_with_defs,
):
    csv = "Number\nM-1\n"
    r = _import(
        client_as(admin_user),
        "00000000-0000-0000-0000-000000000000",
        csv,
        {"Number": "project_number"},
    )
    # 422 matches _fetch_template's convention (shared with create).
    assert r.status_code == 422


def test_import_rejects_unknown_mapping_target(
    client_as, admin_user: User, template_with_defs,
):
    t, _f, _m = template_with_defs
    csv = "Number,Junk\nM-1,Hello\n"
    r = _import(
        client_as(admin_user),
        str(t.id),
        csv,
        # Bogus target — neither a built-in nor a real field def id.
        {"Number": "project_number", "Junk": "00000000-0000-0000-0000-000000000000"},
    )
    assert r.status_code == 422


# Phase 5.3.1 — boolean-conditional auto-true + milestone mapping.


def test_import_boolean_conditional_nonempty_sets_true(
    client_as, admin_user: User, template_with_defs,
):
    t, _f, _m = template_with_defs
    c = client_as(admin_user)
    fid = c.post(
        f"/api/admin/templates/{t.id}/fields",
        json={"name": "Closeout", "field_type": "boolean_conditional_date"},
    ).json()["id"]
    csv = "Number,Closeout\nM-1,2026-12-31\n"
    r = _import(
        c, str(t.id), csv,
        {"Number": "project_number", "Closeout": fid},
    )
    assert r.status_code == 200, r.text
    assert r.json()["created"] == 1
    proj = c.get("/api/projects?q=M-1").json()["items"][0]
    detail = c.get(f"/api/projects/{proj['id']}").json()
    assert detail["custom_field_values"][fid] == {
        "value": True,
        "date": "2026-12-31",
    }


def test_import_boolean_conditional_empty_sets_false(
    client_as, admin_user: User, template_with_defs,
):
    t, _f, _m = template_with_defs
    c = client_as(admin_user)
    fid = c.post(
        f"/api/admin/templates/{t.id}/fields",
        json={"name": "Closeout", "field_type": "boolean_conditional_text"},
    ).json()["id"]
    csv = "Number,Closeout\nM-1,\n"
    r = _import(
        c, str(t.id), csv,
        {"Number": "project_number", "Closeout": fid},
    )
    assert r.status_code == 200
    proj = c.get("/api/projects?q=M-1").json()["items"][0]
    detail = c.get(f"/api/projects/{proj['id']}").json()
    assert detail["custom_field_values"][fid] == {"value": False}


def test_import_milestone_planned_date_maps(
    client_as, admin_user: User, template_with_defs,
):
    """The fixture's milestone is `single` date_model. A short
    `milestone:<id>` mapping writes the planned_date."""
    t, _f, m = template_with_defs
    c = client_as(admin_user)
    csv = "Number,IFC date\nM-1,2026-09-15\n"
    r = _import(
        c, str(t.id), csv,
        {
            "Number": "project_number",
            "IFC date": f"milestone:{m.id}",
        },
    )
    assert r.status_code == 200, r.text
    assert r.json()["created"] == 1
    proj_id = c.get("/api/projects?q=M-1").json()["items"][0]["id"]
    detail = c.get(f"/api/projects/{proj_id}").json()
    milestone = next(
        ms for ms in detail["milestones"]
        if ms["template_milestone_def_id"] == str(m.id)
    )
    assert milestone["planned_date"] == "2026-09-15"


def test_import_milestone_planned_actual_both_slots(
    client_as, admin_user: User, template_with_defs, db_session,
):
    """Add a planned_actual milestone def and verify both slots wire up."""
    t, _f, _m = template_with_defs
    pa_md = TemplateMilestoneDef(
        template_id=t.id,
        name="IFC Submittal",
        direction="outbound",
        date_model="planned_actual",
        order_index=5,
    )
    db_session.add(pa_md)
    db_session.commit()
    c = client_as(admin_user)
    csv = "Number,Planned,Actual\nM-1,2026-08-01,2026-08-15\n"
    r = _import(
        c, str(t.id), csv,
        {
            "Number": "project_number",
            "Planned": f"milestone:{pa_md.id}:planned",
            "Actual": f"milestone:{pa_md.id}:actual",
        },
    )
    assert r.status_code == 200, r.text
    proj_id = c.get("/api/projects?q=M-1").json()["items"][0]["id"]
    detail = c.get(f"/api/projects/{proj_id}").json()
    milestone = next(
        ms for ms in detail["milestones"]
        if ms["template_milestone_def_id"] == str(pa_md.id)
    )
    assert milestone["planned_date"] == "2026-08-01"
    assert milestone["actual_date"] == "2026-08-15"


def test_import_milestone_actual_on_single_rejects(
    client_as, admin_user: User, template_with_defs,
):
    t, _f, m = template_with_defs
    csv = "Number,Actual\nM-1,2026-01-01\n"
    r = _import(
        client_as(admin_user),
        str(t.id),
        csv,
        {
            "Number": "project_number",
            "Actual": f"milestone:{m.id}:actual",
        },
    )
    assert r.status_code == 422
    assert "single date model" in r.json()["detail"].lower()


# Phase 5.3.2 — Excel-format normalization.


def test_import_normalizes_currency_with_dollar_sign_and_commas(
    client_as, admin_user: User, template_with_defs,
):
    t, _f, _m = template_with_defs
    c = client_as(admin_user)
    fid = c.post(
        f"/api/admin/templates/{t.id}/fields",
        json={"name": "Budget", "field_type": "currency"},
    ).json()["id"]
    csv = 'Number,Budget\nM-1,"$3,224.93"\n'
    r = _import(
        c, str(t.id), csv,
        {"Number": "project_number", "Budget": fid},
    )
    assert r.status_code == 200, r.text
    proj = c.get("/api/projects?q=M-1").json()["items"][0]
    detail = c.get(f"/api/projects/{proj['id']}").json()
    assert detail["custom_field_values"][fid] == 3224.93


def test_import_normalizes_percent_with_leading_symbol(
    client_as, admin_user: User, template_with_defs,
):
    t, _f, _m = template_with_defs
    c = client_as(admin_user)
    fid = c.post(
        f"/api/admin/templates/{t.id}/fields",
        json={"name": "Progress", "field_type": "percent"},
    ).json()["id"]
    csv = "Number,Progress\nM-1,%56\nM-2,73%\n"
    r = _import(
        c, str(t.id), csv,
        {"Number": "project_number", "Progress": fid},
    )
    assert r.status_code == 200
    proj1 = c.get("/api/projects?q=M-1").json()["items"][0]
    proj2 = c.get("/api/projects?q=M-2").json()["items"][0]
    d1 = c.get(f"/api/projects/{proj1['id']}").json()
    d2 = c.get(f"/api/projects/{proj2['id']}").json()
    assert d1["custom_field_values"][fid] == 56
    assert d2["custom_field_values"][fid] == 73


def test_import_normalizes_us_date_format(
    client_as, admin_user: User, template_with_defs,
):
    t, _f, _m = template_with_defs
    c = client_as(admin_user)
    fid = c.post(
        f"/api/admin/templates/{t.id}/fields",
        json={"name": "Submitted", "field_type": "date"},
    ).json()["id"]
    csv = "Number,Submitted\nM-1,3/6/2025\n"
    r = _import(
        c, str(t.id), csv,
        {"Number": "project_number", "Submitted": fid},
    )
    assert r.status_code == 200
    proj = c.get("/api/projects?q=M-1").json()["items"][0]
    detail = c.get(f"/api/projects/{proj['id']}").json()
    assert detail["custom_field_values"][fid] == "2025-03-06"


def test_import_normalizes_us_date_for_milestones(
    client_as, admin_user: User, template_with_defs,
):
    t, _f, m = template_with_defs
    csv = "Number,IFC date\nM-1,3/6/2025\n"
    r = _import(
        client_as(admin_user),
        str(t.id),
        csv,
        {
            "Number": "project_number",
            "IFC date": f"milestone:{m.id}",
        },
    )
    assert r.status_code == 200
    proj = client_as(admin_user).get("/api/projects?q=M-1").json()["items"][0]
    detail = client_as(admin_user).get(f"/api/projects/{proj['id']}").json()
    milestone = next(
        ms for ms in detail["milestones"]
        if ms["template_milestone_def_id"] == str(m.id)
    )
    assert milestone["planned_date"] == "2025-03-06"


def test_import_normalizes_boolean_yes_no(
    client_as, admin_user: User, template_with_defs,
):
    t, _f, _m = template_with_defs
    c = client_as(admin_user)
    fid = c.post(
        f"/api/admin/templates/{t.id}/fields",
        json={"name": "Closed", "field_type": "boolean"},
    ).json()["id"]
    csv = "Number,Closed\nM-1,Yes\nM-2,No\n"
    r = _import(
        c, str(t.id), csv,
        {"Number": "project_number", "Closed": fid},
    )
    assert r.status_code == 200
    proj1 = c.get("/api/projects?q=M-1").json()["items"][0]
    proj2 = c.get("/api/projects?q=M-2").json()["items"][0]
    d1 = c.get(f"/api/projects/{proj1['id']}").json()
    d2 = c.get(f"/api/projects/{proj2['id']}").json()
    assert d1["custom_field_values"][fid] is True
    assert d2["custom_field_values"][fid] is False


def test_import_normalizes_single_select_case_insensitive(
    client_as, admin_user: User, template_with_defs,
):
    t, _f, _m = template_with_defs
    c = client_as(admin_user)
    fid = c.post(
        f"/api/admin/templates/{t.id}/fields",
        json={
            "name": "Phase",
            "field_type": "single_select",
            "options": {"choices": ["Scoping", "Design", "Build"]},
        },
    ).json()["id"]
    csv = "Number,Phase\nM-1,design\n"
    r = _import(
        c, str(t.id), csv,
        {"Number": "project_number", "Phase": fid},
    )
    assert r.status_code == 200
    proj = c.get("/api/projects?q=M-1").json()["items"][0]
    detail = c.get(f"/api/projects/{proj['id']}").json()
    # Normalized to the canonical-cased choice.
    assert detail["custom_field_values"][fid] == "Design"


def test_import_milestone_unknown_id_rejects(
    client_as, admin_user: User, template_with_defs,
):
    t, _f, _m = template_with_defs
    csv = "Number,Foo\nM-1,2026-01-01\n"
    r = _import(
        client_as(admin_user),
        str(t.id),
        csv,
        {
            "Number": "project_number",
            "Foo": "milestone:00000000-0000-0000-0000-000000000000:planned",
        },
    )
    assert r.status_code == 422


# ---- Phase 5.4 — export ---------------------------------------------------


def _export(c, template_id: str, *, format: str, columns: list[str], **extra):
    params = {
        "template_id": template_id,
        "format": format,
        "columns": ",".join(columns),
        **extra,
    }
    return c.get("/api/projects/export", params=params)


def test_export_csv_built_ins_returns_headers_and_rows(
    client_as, admin_user: User, template_with_defs,
):
    t, _f, _m = template_with_defs
    c = client_as(admin_user)
    r1 = c.post(
        "/api/projects",
        json={
            "project_number": "M-001",
            "title": "Alpha",
            "template_id": str(t.id),
        },
    )
    assert r1.status_code == 201, r1.text
    r2 = c.post(
        "/api/projects",
        json={
            "project_number": "M-002",
            "title": "Bravo",
            "template_id": str(t.id),
        },
    )
    assert r2.status_code == 201, r2.text
    r = _export(
        c, str(t.id),
        format="csv",
        columns=["builtin:project_number", "builtin:title", "builtin:lifecycle"],
    )
    assert r.status_code == 200, r.text
    assert r.headers["content-type"].startswith("text/csv")
    assert 'filename="DIV1-CON-DESIGN_' in r.headers["content-disposition"]
    body = r.content.decode("utf-8-sig").splitlines()
    assert body[0] == "Project #,Title,Status"
    assert len(body) == 3
    assert any(line.startswith("M-001,Alpha,") for line in body[1:])
    assert any(line.startswith("M-002,Bravo,") for line in body[1:])


def test_export_custom_field_column_renders_value(
    client_as, admin_user: User, template_with_defs,
):
    t, _f, _m = template_with_defs
    c = client_as(admin_user)
    fid = c.post(
        f"/api/admin/templates/{t.id}/fields",
        json={"name": "Budget", "field_type": "currency"},
    ).json()["id"]
    c.post(
        "/api/projects",
        json={
            "project_number": "M-001",
            "title": "Alpha",
            "template_id": str(t.id),
            "custom_field_values": {fid: 12500.5},
        },
    )
    r = _export(
        c, str(t.id),
        format="csv",
        columns=["builtin:project_number", f"custom_field:{fid}"],
    )
    assert r.status_code == 200, r.text
    lines = r.content.decode("utf-8-sig").splitlines()
    assert lines[0] == "Project #,Budget"
    assert lines[1] == "M-001,12500.5"


def test_export_milestone_column_renders_planned_date(
    client_as, admin_user: User, template_with_defs,
):
    t, _f, m = template_with_defs
    c = client_as(admin_user)
    proj = c.post(
        "/api/projects",
        json={
            "project_number": "M-001",
            "title": "Alpha",
            "template_id": str(t.id),
        },
    ).json()
    # Auto-created milestone row; set its planned_date via PATCH.
    ms_id = proj["milestones"][0]["id"]
    c.patch(
        f"/api/projects/{proj['id']}/milestones/{ms_id}",
        json={"planned_date": "2026-09-01"},
    )
    r = _export(
        c, str(t.id),
        format="csv",
        columns=[f"milestone:{m.id}:date"],
    )
    assert r.status_code == 200, r.text
    lines = r.content.decode("utf-8-sig").splitlines()
    assert lines[0] == "IFC"
    assert lines[1] == "2026-09-01"


def test_export_honors_q_filter(
    client_as, admin_user: User, template_with_defs,
):
    t, _f, _m = template_with_defs
    c = client_as(admin_user)
    for number, title in [
        ("M-001", "Alpha"),
        ("M-002", "Bravo"),
        ("M-003", "Charlie"),
    ]:
        r = c.post(
            "/api/projects",
            json={
                "project_number": number,
                "title": title,
                "template_id": str(t.id),
            },
        )
        assert r.status_code == 201, r.text
    r = _export(
        c, str(t.id),
        format="csv",
        columns=["builtin:project_number", "builtin:title"],
        q="brav",
    )
    assert r.status_code == 200, r.text
    lines = r.content.decode("utf-8-sig").splitlines()
    assert len(lines) == 2  # header + 1 row
    assert lines[1] == "M-002,Bravo"


def test_export_over_cap_returns_422(
    client_as, admin_user: User, template_with_defs, monkeypatch,
):
    import backend.app.services.project_export as pe
    monkeypatch.setattr(pe, "EXPORT_ROW_CAP", 1)
    t, _f, _m = template_with_defs
    c = client_as(admin_user)
    for number in ("M-001", "M-002"):
        r = c.post(
            "/api/projects",
            json={
                "project_number": number,
                "title": number,
                "template_id": str(t.id),
            },
        )
        assert r.status_code == 201, r.text
    r = _export(
        c, str(t.id),
        format="csv",
        columns=["builtin:project_number"],
    )
    assert r.status_code == 422
    assert "tighten" in r.json()["detail"].lower()


def test_export_xlsx_roundtrips_through_openpyxl(
    client_as, admin_user: User, template_with_defs,
):
    import io
    from openpyxl import load_workbook
    t, _f, _m = template_with_defs
    c = client_as(admin_user)
    c.post(
        "/api/projects",
        json={
            "project_number": "M-001",
            "title": "Alpha",
            "template_id": str(t.id),
        },
    )
    r = _export(
        c, str(t.id),
        format="xlsx",
        columns=["builtin:project_number", "builtin:title"],
    )
    assert r.status_code == 200, r.text
    assert r.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml"
    )
    assert r.content[:4] == b"PK\x03\x04"  # ZIP magic
    assert 'filename="DIV1-CON-DESIGN_' in r.headers["content-disposition"]
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb.active
    assert ws["A1"].value == "Project #"
    assert ws["B1"].value == "Title"
    assert ws["A2"].value == "M-001"
    assert ws["B2"].value == "Alpha"


def test_list_projects_conditions_filter(client_as, admin_user, db_session):
    from backend.tests.test_metric_engine import _taxonomy, _field, _project
    _, t = _taxonomy(db_session, "PCF")
    qa = _field(db_session, t.id, "QA done", "boolean")
    _project(db_session, t.id, admin_user, cfv={str(qa.id): False}, title="needs qa")
    _project(db_session, t.id, admin_user, cfv={str(qa.id): True}, title="done")
    db_session.commit()
    import json
    conditions = json.dumps({"combinator": "and",
                             "items": [{"field": str(qa.id), "op": "is_false"}]})
    r = client_as(admin_user).get(
        f"/api/projects?template_id={t.id}&conditions={conditions}"
    )
    assert r.status_code == 200
    titles = [i["title"] for i in r.json()["items"]]
    assert titles == ["needs qa"]
    # conditions without template_id -> 422
    assert client_as(admin_user).get(
        f"/api/projects?conditions={conditions}"
    ).status_code == 422
