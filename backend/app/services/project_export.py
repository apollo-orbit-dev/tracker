"""Row + cell rendering for Saved-View CSV / XLSX exports (Phase 5.4).

Owns the column-key parsing and per-cell value extraction. Returns
typed Python values so the XLSX renderer can keep numeric/date cells
typed; the CSV renderer stringifies everything in one place.
"""
import csv
import io
import re
import uuid
from datetime import date, datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from backend.app.db.models import (
    Milestone,
    Project,
    TemplateFieldDef,
    TemplateMilestoneDef,
)
from backend.app.services.ref_labels import collect_ref_labels
from backend.app.services.view_columns import parse_column_key

EXPORT_ROW_CAP = 5000

_BUILTIN_LABELS = {
    "project_number": "Project #",
    "client_number": "Client #",
    "title": "Title",
    "lifecycle": "Status",
    "created_at": "Created",
    "updated_at": "Updated",
}

# CSV / XLSX value resolver: returns (value, field_type) where field_type
# drives the XLSX cell format. CSV stringifies; XLSX applies number/date
# formats per type. Unknown / removed entities return ("—", "text") so
# the row stays aligned but doesn't render a misleading value.

_NUMERIC_TYPES = {"integer", "decimal", "currency", "percent", "auto_number"}


class ColumnsError(Exception):
    """Raised when a column key is malformed or refers to a missing field/milestone."""

    def __init__(self, message: str) -> None:
        self.message = message
        super().__init__(message)


def validate_export_columns(
    columns: list[str],
    *,
    live_custom_field_ids: set[uuid.UUID],
    live_milestone_def_ids: set[uuid.UUID],
) -> list[tuple[str, str, str | None]]:
    """Parse + validate column keys against this template's live defs.

    Returns the parsed shape: list of (category, ident, mode) tuples,
    preserving caller order. Raises ColumnsError on the first bad key
    so the API returns a clear 422.
    """
    if not columns:
        raise ColumnsError("columns is required")
    parsed_keys: list[tuple[str, str, str | None]] = []
    seen: set[str] = set()
    for key in columns:
        if key in seen:
            raise ColumnsError(f"duplicate column key: {key}")
        seen.add(key)
        parsed = parse_column_key(key)
        if parsed is None:
            raise ColumnsError(f"invalid column key: {key}")
        category, ident, mode = parsed
        if category == "custom_field":
            if uuid.UUID(ident) not in live_custom_field_ids:
                raise ColumnsError(f"custom field not in this template: {key}")
        elif category == "milestone":
            if uuid.UUID(ident) not in live_milestone_def_ids:
                raise ColumnsError(f"milestone not in this template: {key}")
        parsed_keys.append((category, ident, mode))
    return parsed_keys


def header_label(
    parsed_key: tuple[str, str, str | None],
    field_defs_by_id: dict[str, TemplateFieldDef],
    milestone_defs_by_id: dict[str, TemplateMilestoneDef],
) -> str:
    category, ident, mode = parsed_key
    if category == "builtin":
        return _BUILTIN_LABELS.get(ident, ident)
    if category == "custom_field":
        fd = field_defs_by_id.get(ident)
        return fd.name if fd is not None else "(removed field)"
    md = milestone_defs_by_id.get(ident)
    if md is None:
        return "(removed milestone)"
    if mode == "date":
        return md.name
    return f"{md.name} ({mode})"


def _milestone_value(
    project: Project,
    milestone_def_id: str,
    mode: str | None,
) -> date | None:
    target = uuid.UUID(milestone_def_id)
    for ms in project.milestones:
        if ms.deleted_at is not None:
            continue
        if ms.template_milestone_def_id != target:
            continue
        if mode == "actual":
            return ms.actual_date
        return ms.planned_date
    return None


def _cf_value(
    project: Project,
    field_def: TemplateFieldDef | None,
    ref_labels: dict[str, dict[str, str]],
) -> tuple[Any, str]:
    """Return (rendered_value, field_type_for_format)."""
    if field_def is None:
        return ("(removed)", "text")
    raw = (project.custom_field_values or {}).get(str(field_def.id))
    ftype = field_def.field_type
    if raw is None or raw == "":
        return (None, ftype)
    if ftype == "user_picker_single":
        return (ref_labels.get("users", {}).get(str(raw), str(raw)), "text")
    if ftype == "user_picker_multi":
        if not isinstance(raw, list):
            return (str(raw), "text")
        names = [ref_labels.get("users", {}).get(str(u), str(u)) for u in raw]
        return (", ".join(names), "text")
    if ftype == "contact_picker":
        return (ref_labels.get("contacts", {}).get(str(raw), str(raw)), "text")
    if ftype == "project_reference":
        return (ref_labels.get("projects", {}).get(str(raw), str(raw)), "text")
    if ftype == "client_reference":
        return (ref_labels.get("clients", {}).get(str(raw), str(raw)), "text")
    if ftype == "date":
        if isinstance(raw, str):
            try:
                return (date.fromisoformat(raw), "date")
            except ValueError:
                return (raw, "text")
        return (raw, "date")
    if ftype == "boolean":
        return (bool(raw), "boolean")
    if ftype in ("boolean_conditional_date", "boolean_conditional_text"):
        if not isinstance(raw, dict):
            return (str(raw), "text")
        if not raw.get("value"):
            return (False, "boolean")
        if ftype == "boolean_conditional_date":
            inner = raw.get("date")
            if isinstance(inner, str):
                try:
                    return (date.fromisoformat(inner), "date")
                except ValueError:
                    return (inner, "text")
            return (inner, "date")
        return (raw.get("text") or "", "text")
    if ftype == "multi_select":
        if isinstance(raw, list):
            return (", ".join(str(v) for v in raw), "text")
        return (str(raw), "text")
    if ftype in _NUMERIC_TYPES:
        try:
            return (float(raw) if ftype != "integer" and ftype != "auto_number" else int(raw), ftype)
        except (TypeError, ValueError):
            return (str(raw), "text")
    # short_text, long_text, single_select, anything else
    return (str(raw), "text")


def _builtin_value(
    project: Project, name: str
) -> tuple[Any, str]:
    if name == "project_number":
        return (project.project_number, "text")
    if name == "client_number":
        return (project.client_project_number, "text")
    if name == "title":
        return (project.title, "text")
    if name == "lifecycle":
        return (project.lifecycle_state, "text")
    if name == "created_at":
        return (project.created_at, "datetime")
    if name == "updated_at":
        return (project.updated_at, "datetime")
    return ("", "text")


def render_row(
    project: Project,
    parsed_keys: list[tuple[str, str, str | None]],
    field_defs_by_id: dict[str, TemplateFieldDef],
    ref_labels: dict[str, dict[str, str]],
) -> list[tuple[Any, str]]:
    """Render one project as a list of (value, type_for_format) cells.

    `type_for_format` is one of: text, integer, decimal, currency,
    percent, auto_number, date, datetime, boolean.
    """
    out: list[tuple[Any, str]] = []
    for parsed in parsed_keys:
        category, ident, mode = parsed
        if category == "builtin":
            out.append(_builtin_value(project, ident))
        elif category == "custom_field":
            fd = field_defs_by_id.get(ident)
            out.append(_cf_value(project, fd, ref_labels))
        else:  # milestone
            d = _milestone_value(project, ident, mode)
            out.append((d, "date"))
    return out


def _stringify_csv(value: Any, ftype: str) -> str:
    if value is None:
        return ""
    if ftype == "datetime":
        if isinstance(value, datetime):
            return value.isoformat(timespec="seconds")
        return str(value)
    if ftype == "date":
        if isinstance(value, date) and not isinstance(value, datetime):
            return value.isoformat()
        if isinstance(value, datetime):
            return value.date().isoformat()
        return str(value)
    if ftype == "boolean":
        return "TRUE" if value else "FALSE"
    if ftype == "percent":
        # Stored as 0-100; render with a % suffix so Excel users see
        # what they expect when reopening a CSV.
        return f"{value}%"
    return str(value)


def render_csv(
    headers: list[str],
    rows: list[list[tuple[Any, str]]],
) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    for row in rows:
        writer.writerow([_stringify_csv(v, t) for v, t in row])
    return buf.getvalue().encode("utf-8-sig")


def _xlsx_format_for(ftype: str) -> str | None:
    if ftype == "currency":
        return "$#,##0.00"
    if ftype == "decimal":
        return "#,##0.00"
    if ftype in ("integer", "auto_number"):
        return "#,##0"
    if ftype == "percent":
        # Stored 0-100, so divide by 100 in render or use a 0\% style.
        # We use the explicit "0\%" style and write the raw number so
        # Excel shows e.g. 56% without dividing.
        return '0"%"'
    if ftype == "date":
        return "yyyy-mm-dd"
    if ftype == "datetime":
        return "yyyy-mm-dd hh:mm:ss"
    return None


_SHEET_NAME_BAD = re.compile(r'[\\/?*:\[\]]')


def _sheet_name(intersection: str) -> str:
    cleaned = _SHEET_NAME_BAD.sub("-", intersection).strip()
    if not cleaned:
        cleaned = "Projects"
    return cleaned[:31]


def render_xlsx(
    headers: list[str],
    rows: list[list[tuple[Any, str]]],
    sheet_name: str,
) -> bytes:
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title=_sheet_name(sheet_name))
    ws.append(headers)
    # Track per-column format and width.
    col_formats: dict[int, str] = {}
    for row in rows:
        cells: list[Any] = []
        for col_idx, (value, ftype) in enumerate(row, start=1):
            fmt = _xlsx_format_for(ftype)
            if fmt is not None and col_idx not in col_formats:
                col_formats[col_idx] = fmt
            if value is None:
                cells.append(None)
                continue
            if ftype == "datetime" and isinstance(value, datetime):
                cells.append(value.replace(tzinfo=None))
            elif ftype == "boolean":
                cells.append(bool(value))
            else:
                cells.append(value)
        ws.append(cells)
    # Apply formats. Write-only mode requires setting via column dimensions.
    for col_idx, fmt in col_formats.items():
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].number_format = fmt
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


_SLUG_BAD = re.compile(r"[^A-Z0-9]+")


def slug_for_filename(parts: list[str], today: date) -> str:
    """Build the filename stem: 'DIV1-CON-DESIGN_2026-06-09'.

    Each part is uppercased and stripped down to A-Z0-9 (so `Design` →
    `PC`, not `P-C`). The hyphen is reserved as the part separator.
    """
    cleaned = []
    for p in parts:
        s = _SLUG_BAD.sub("", (p or "").upper())
        if s:
            cleaned.append(s)
    stem = "-".join(cleaned) if cleaned else "PROJECTS"
    return f"{stem}_{today.isoformat()}"


def gather_export_context(
    db,
    *,
    template,
    projects: list[Project],
):
    """Pre-load field defs + milestone defs + ref labels in bulk for an
    export. Returns (field_defs_by_id, milestone_defs_by_id, ref_labels).
    """
    from sqlalchemy import select
    field_defs = list(
        db.execute(
            select(TemplateFieldDef).where(
                TemplateFieldDef.template_id == template.id,
                TemplateFieldDef.deleted_at.is_(None),
            )
        ).scalars()
    )
    milestone_defs = list(
        db.execute(
            select(TemplateMilestoneDef).where(
                TemplateMilestoneDef.template_id == template.id,
                TemplateMilestoneDef.deleted_at.is_(None),
            )
        ).scalars()
    )
    field_defs_by_id = {str(fd.id): fd for fd in field_defs}
    milestone_defs_by_id = {str(md.id): md for md in milestone_defs}
    # Milestones for each project: eager-loaded by the caller; safety
    # fallback if not — fetch in one query.
    if projects and not all(
        ("milestones" in p.__dict__) for p in projects
    ):
        pids = [p.id for p in projects]
        all_ms = list(
            db.execute(
                select(Milestone).where(Milestone.project_id.in_(pids))
            ).scalars()
        )
        by_proj: dict[uuid.UUID, list[Milestone]] = {pid: [] for pid in pids}
        for m in all_ms:
            by_proj[m.project_id].append(m)
        for p in projects:
            p.milestones = by_proj.get(p.id, [])  # type: ignore[assignment]
    ref_labels = collect_ref_labels(
        db, projects=projects, live_field_defs=field_defs
    )
    return field_defs_by_id, milestone_defs_by_id, ref_labels
